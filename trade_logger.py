"""
╔══════════════════════════════════════════════════════════════╗
║   POLYMARKET TELEGRAM BOT — EXCEL TRADE LOGGER              ║
╚══════════════════════════════════════════════════════════════╝
Logs all trading activity to an Excel (.xlsx) file.

Sheets:
  1. Trades   — Every trade with entry/exit details, P&L
  2. Events   — Activity log (bot start, TP/SL, redeem, errors)
"""

import os
import threading
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# File path for the Excel log
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_log.xlsx")

# Column headers
TRADE_HEADERS = [
    "Trade #", "Date", "Time", "Direction", "Market", "Timeframe",
    "Entry Price", "Exit Price", "Shares", "Stake ($)",
    "P&L ($)", "P&L (%)", "Close Reason", "Order ID", "Duration (min)"
]

EVENT_HEADERS = [
    "Date", "Time", "Event Type", "Details"
]

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


class TradeLogger:
    """Logs all trades and events to an Excel file."""

    def __init__(self, file_path: str = None):
        self.file_path = file_path or LOG_FILE
        self._lock = threading.Lock()
        self._trade_count = 0
        self._ensure_workbook()

    def _ensure_workbook(self):
        """Create the workbook if it doesn't exist, or load existing."""
        if os.path.exists(self.file_path):
            try:
                wb = load_workbook(self.file_path)
                # Count existing trades
                if "Trades" in wb.sheetnames:
                    self._trade_count = wb["Trades"].max_row - 1  # minus header
                    if self._trade_count < 0:
                        self._trade_count = 0
                wb.close()
                return
            except Exception:
                pass  # File corrupted, recreate

        # Create new workbook
        wb = Workbook()

        # Trades sheet
        ws_trades = wb.active
        ws_trades.title = "Trades"
        ws_trades.append(TRADE_HEADERS)
        self._style_headers(ws_trades, len(TRADE_HEADERS))
        self._set_column_widths(ws_trades, [8, 12, 10, 10, 45, 10, 12, 12, 10, 10, 10, 10, 20, 25, 12])

        # Events sheet
        ws_events = wb.create_sheet("Events")
        ws_events.append(EVENT_HEADERS)
        self._style_headers(ws_events, len(EVENT_HEADERS))
        self._set_column_widths(ws_events, [12, 10, 18, 80])

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Value"])
        self._style_headers(ws_summary, 2)
        ws_summary.append(["Total Trades", 0])
        ws_summary.append(["Winning Trades", 0])
        ws_summary.append(["Losing Trades", 0])
        ws_summary.append(["Win Rate (%)", "0.0%"])
        ws_summary.append(["Total P&L ($)", 0])
        ws_summary.append(["Best Trade ($)", 0])
        ws_summary.append(["Worst Trade ($)", 0])
        ws_summary.append(["Avg P&L per Trade ($)", 0])
        ws_summary.append(["Total Volume ($)", 0])
        ws_summary.append(["First Trade", "N/A"])
        ws_summary.append(["Last Trade", "N/A"])
        self._set_column_widths(ws_summary, [22, 20])

        wb.save(self.file_path)
        wb.close()

    def _style_headers(self, ws, col_count):
        """Apply styles to header row."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def _set_column_widths(self, ws, widths):
        """Set column widths."""
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def log_trade(self, trade) -> None:
        """Log a completed trade to the Trades sheet."""
        threading.Thread(target=self._write_trade, args=(trade,), daemon=True).start()

    def _write_trade(self, trade):
        """Write trade data to Excel (thread-safe)."""
        with self._lock:
            try:
                wb = load_workbook(self.file_path)
                ws = wb["Trades"]
                self._trade_count += 1

                now = datetime.now()
                direction = trade.direction.value if hasattr(trade.direction, 'value') else str(trade.direction)

                # Calculate duration
                duration_min = 0
                if hasattr(trade, 'entry_time') and hasattr(trade, 'close_time'):
                    try:
                        if trade.close_time and trade.entry_time:
                            entry_dt = datetime.strptime(trade.entry_time, "%Y-%m-%d %H:%M:%S")
                            close_dt = datetime.strptime(trade.close_time, "%Y-%m-%d %H:%M:%S")
                            duration_min = (close_dt - entry_dt).total_seconds() / 60
                    except Exception:
                        pass

                # P&L percentage
                pnl_pct = 0
                if trade.amount > 0:
                    pnl_pct = (trade.pnl / trade.amount) * 100

                row_data = [
                    self._trade_count,
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S"),
                    direction,
                    getattr(trade, 'market_question', '')[:45] or trade.timeframe,
                    getattr(trade, 'timeframe', 'N/A'),
                    round(trade.share_price, 4),
                    round(getattr(trade, 'result_price', trade.current_price or 0), 4),
                    round(trade.shares, 2),
                    round(trade.amount, 2),
                    round(trade.pnl, 2),
                    f"{pnl_pct:.1f}%",
                    trade.close_reason or "Unknown",
                    trade.order_id[:20] if trade.order_id else "N/A",
                    round(duration_min, 1),
                ]

                ws.append(row_data)
                row_num = ws.max_row

                # Color row based on P&L
                fill = GREEN_FILL if trade.pnl >= 0 else RED_FILL
                for col in range(1, len(TRADE_HEADERS) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="center")

                # Update summary
                self._update_summary(wb)

                wb.save(self.file_path)
                wb.close()
            except Exception as e:
                print(f"Trade log error: {e}")

    def log_event(self, event_type: str, details: str) -> None:
        """Log an event to the Events sheet."""
        threading.Thread(target=self._write_event, args=(event_type, details), daemon=True).start()

    def _write_event(self, event_type: str, details: str):
        """Write event data to Excel (thread-safe)."""
        with self._lock:
            try:
                wb = load_workbook(self.file_path)
                ws = wb["Events"]
                now = datetime.now()
                ws.append([
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S"),
                    event_type,
                    details[:200],
                ])
                row_num = ws.max_row
                for col in range(1, len(EVENT_HEADERS) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = BORDER

                wb.save(self.file_path)
                wb.close()
            except Exception as e:
                print(f"Event log error: {e}")

    def _update_summary(self, wb):
        """Update the Summary sheet with latest stats."""
        try:
            ws_trades = wb["Trades"]
            ws_summary = wb["Summary"]

            # Collect all P&L values
            pnls = []
            volumes = []
            for row in ws_trades.iter_rows(min_row=2, values_only=True):
                if row[10] is not None:  # P&L column
                    try:
                        pnl = float(row[10])
                        pnls.append(pnl)
                    except (ValueError, TypeError):
                        pass
                if row[9] is not None:  # Stake column
                    try:
                        volumes.append(float(row[9]))
                    except (ValueError, TypeError):
                        pass

            total = len(pnls)
            wins = sum(1 for p in pnls if p >= 0)
            losses = sum(1 for p in pnls if p < 0)
            win_rate = (wins / total * 100) if total > 0 else 0
            total_pnl = sum(pnls)
            best = max(pnls) if pnls else 0
            worst = min(pnls) if pnls else 0
            avg_pnl = total_pnl / total if total > 0 else 0
            total_vol = sum(volumes)

            # First and last trade dates
            first_date = ws_trades.cell(row=2, column=2).value if total > 0 else "N/A"
            last_date = ws_trades.cell(row=ws_trades.max_row, column=2).value if total > 0 else "N/A"

            # Write summary
            summary_data = [
                total, wins, losses, f"{win_rate:.1f}%",
                round(total_pnl, 2), round(best, 2), round(worst, 2),
                round(avg_pnl, 2), round(total_vol, 2),
                first_date, last_date,
            ]
            for i, val in enumerate(summary_data, 2):
                ws_summary.cell(row=i, column=2, value=val)

        except Exception:
            pass

    def get_history(self, period: str = None) -> dict:
        """
        Get trade history filtered by period.
        
        period formats:
            None / "all"    → all time
            "2024"          → year 2024
            "2024-03"       → March 2024
            "today"         → today only
            "1d" / "7d"     → last N days
            "1w" / "2w"     → last N weeks
            "1m" / "3m"     → last N months
            "1y" / "2y"     → last N years
        
        Returns dict with: trades (list), stats (dict)
        """
        from datetime import datetime, timedelta

        if not os.path.exists(self.file_path):
            return {"trades": [], "stats": {}, "period_label": "No data"}

        with self._lock:
            try:
                wb = load_workbook(self.file_path, read_only=True)
                ws = wb["Trades"]

                # Read all trades
                all_trades = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    all_trades.append({
                        "num": row[0],
                        "date": str(row[1]) if row[1] else "",
                        "time": str(row[2]) if row[2] else "",
                        "direction": row[3] or "",
                        "market": row[4] or "",
                        "timeframe": row[5] or "",
                        "entry_price": row[6] or 0,
                        "exit_price": row[7] or 0,
                        "shares": row[8] or 0,
                        "stake": row[9] or 0,
                        "pnl": row[10] or 0,
                        "pnl_pct": row[11] or "0%",
                        "close_reason": row[12] or "",
                        "order_id": row[13] or "",
                        "duration": row[14] or 0,
                    })
                wb.close()

                # Determine date filter
                now = datetime.now()
                start_date = None
                period_label = "All Time"

                if period and period.lower() != "all":
                    p = period.lower().strip()

                    if p == "today":
                        start_date = now.replace(hour=0, minute=0, second=0)
                        period_label = f"Today ({now.strftime('%Y-%m-%d')})"

                    elif p.endswith("d") and p[:-1].isdigit():
                        days = int(p[:-1])
                        start_date = now - timedelta(days=days)
                        period_label = f"Last {days} day(s)"

                    elif p.endswith("w") and p[:-1].isdigit():
                        weeks = int(p[:-1])
                        start_date = now - timedelta(weeks=weeks)
                        period_label = f"Last {weeks} week(s)"

                    elif p.endswith("m") and p[:-1].isdigit():
                        months = int(p[:-1])
                        start_date = now - timedelta(days=months * 30)
                        period_label = f"Last {months} month(s)"

                    elif p.endswith("y") and p[:-1].isdigit():
                        years = int(p[:-1])
                        start_date = now - timedelta(days=years * 365)
                        period_label = f"Last {years} year(s)"

                    elif len(p) == 4 and p.isdigit():
                        # Year only: "2024"
                        year = int(p)
                        period_label = f"Year {year}"
                        # Filter by year string match
                        all_trades = [t for t in all_trades if t["date"].startswith(p)]
                        start_date = None  # Already filtered

                    elif len(p) == 7 and p[4] == "-":
                        # Year-month: "2024-03"
                        period_label = f"Month {p}"
                        all_trades = [t for t in all_trades if t["date"].startswith(p)]
                        start_date = None

                # Apply date filter if start_date is set
                if start_date:
                    start_str = start_date.strftime("%Y-%m-%d")
                    all_trades = [t for t in all_trades if t["date"] >= start_str]

                # Calculate stats
                pnls = []
                volumes = []
                for t in all_trades:
                    try:
                        pnl = float(t["pnl"])
                        pnls.append(pnl)
                    except (ValueError, TypeError):
                        pass
                    try:
                        volumes.append(float(t["stake"]))
                    except (ValueError, TypeError):
                        pass

                total = len(pnls)
                wins = sum(1 for p in pnls if p >= 0)
                losses = sum(1 for p in pnls if p < 0)
                win_rate = (wins / total * 100) if total > 0 else 0
                total_pnl = sum(pnls)
                best = max(pnls) if pnls else 0
                worst = min(pnls) if pnls else 0
                avg_pnl = total_pnl / total if total > 0 else 0
                total_vol = sum(volumes)

                stats = {
                    "total": total,
                    "wins": wins,
                    "losses": losses,
                    "win_rate": win_rate,
                    "total_pnl": total_pnl,
                    "best": best,
                    "worst": worst,
                    "avg_pnl": avg_pnl,
                    "total_volume": total_vol,
                    "first_date": all_trades[0]["date"] if all_trades else "N/A",
                    "last_date": all_trades[-1]["date"] if all_trades else "N/A",
                }

                return {
                    "trades": all_trades,
                    "stats": stats,
                    "period_label": period_label,
                }

            except Exception as e:
                return {"trades": [], "stats": {}, "period_label": f"Error: {str(e)[:100]}"}

