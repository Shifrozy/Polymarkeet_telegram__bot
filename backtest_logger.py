"""
╔══════════════════════════════════════════════════════════════╗
║   POLYMARKET TELEGRAM BOT — BACKTEST LOGGER                  ║
╚══════════════════════════════════════════════════════════════╝
Separate Excel file for backtesting / simulation data.
File: backtest_log.xlsx

Sheets:
  1. Trades    — All backtested trades with P&L
  2. Events    — Backtest run logs
  3. Summary   — Auto-calculated stats
  4. Daily     — Daily P&L breakdown
"""

import os
import threading
from datetime import datetime
from typing import Optional, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# File path
BACKTEST_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backtest_log.xlsx")

# Column headers
TRADE_HEADERS = [
    "Trade #", "Date", "Time", "Direction", "Market", "Timeframe",
    "Entry Price", "Exit Price", "Shares", "Stake ($)",
    "P&L ($)", "P&L (%)", "Close Reason", "Duration (min)", "Balance ($)"
]

EVENT_HEADERS = [
    "Date", "Time", "Event Type", "Details"
]

DAILY_HEADERS = [
    "Date", "Trades", "Wins", "Losses", "Win Rate (%)",
    "P&L ($)", "Volume ($)", "Balance ($)"
]

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


class BacktestLogger:
    """Logs backtesting data to a separate Excel file."""

    def __init__(self, file_path: str = None, starting_balance: float = 100.0):
        self.file_path = file_path or BACKTEST_FILE
        self._lock = threading.Lock()
        self._trade_count = 0
        self._balance = starting_balance
        self._starting_balance = starting_balance
        self._daily_data = {}  # {date_str: {trades, wins, losses, pnl, volume}}
        self._create_fresh()

    def _create_fresh(self):
        """Create a fresh backtest workbook (overwrites existing)."""
        wb = Workbook()

        # Trades sheet
        ws_trades = wb.active
        ws_trades.title = "Trades"
        ws_trades.append(TRADE_HEADERS)
        self._style_headers(ws_trades, len(TRADE_HEADERS))
        self._set_column_widths(ws_trades, [8, 12, 10, 10, 45, 10, 12, 12, 10, 10, 10, 10, 18, 12, 12])

        # Events sheet
        ws_events = wb.create_sheet("Events")
        ws_events.append(EVENT_HEADERS)
        self._style_headers(ws_events, len(EVENT_HEADERS))
        self._set_column_widths(ws_events, [12, 10, 18, 80])

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Value"])
        self._style_headers(ws_summary, 2)
        summary_rows = [
            ["Starting Balance ($)", self._starting_balance],
            ["Final Balance ($)", self._starting_balance],
            ["Total P&L ($)", 0],
            ["Return (%)", "0.0%"],
            ["Total Trades", 0],
            ["Winning Trades", 0],
            ["Losing Trades", 0],
            ["Win Rate (%)", "0.0%"],
            ["Best Trade ($)", 0],
            ["Worst Trade ($)", 0],
            ["Avg P&L per Trade ($)", 0],
            ["Max Drawdown ($)", 0],
            ["Max Drawdown (%)", "0.0%"],
            ["Profit Factor", "0.0"],
            ["Total Volume ($)", 0],
            ["Avg Trade Duration (min)", 0],
            ["First Trade", "N/A"],
            ["Last Trade", "N/A"],
            ["Backtest Run", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        for row in summary_rows:
            ws_summary.append(row)
        self._set_column_widths(ws_summary, [24, 22])

        # Daily sheet
        ws_daily = wb.create_sheet("Daily")
        ws_daily.append(DAILY_HEADERS)
        self._style_headers(ws_daily, len(DAILY_HEADERS))
        self._set_column_widths(ws_daily, [12, 8, 8, 8, 12, 10, 12, 12])

        # Write BACKTEST_START event directly (not via _write_event to avoid lock)
        ws_events = wb["Events"]
        now = datetime.now()
        ws_events.append([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            "BACKTEST_START",
            f"Starting balance: ${self._starting_balance:.2f}",
        ])

        wb.save(self.file_path)
        wb.close()

    def _style_headers(self, ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def _set_column_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def log_trade(self, trade_data: dict) -> None:
        """
        Log a backtest trade.
        trade_data keys: direction, market, timeframe, entry_price, exit_price,
                         shares, stake, pnl, pnl_pct, close_reason, duration,
                         date (optional), time (optional)
        """
        with self._lock:
            try:
                wb = load_workbook(self.file_path)
                ws = wb["Trades"]
                self._trade_count += 1
                self._balance += trade_data.get("pnl", 0)

                now = datetime.now()
                date_str = trade_data.get("date", now.strftime("%Y-%m-%d"))
                time_str = trade_data.get("time", now.strftime("%H:%M:%S"))

                pnl = trade_data.get("pnl", 0)
                stake = trade_data.get("stake", 0)
                pnl_pct = (pnl / stake * 100) if stake > 0 else 0

                row_data = [
                    self._trade_count,
                    date_str,
                    time_str,
                    trade_data.get("direction", ""),
                    trade_data.get("market", "")[:45],
                    trade_data.get("timeframe", "15m"),
                    round(trade_data.get("entry_price", 0), 4),
                    round(trade_data.get("exit_price", 0), 4),
                    round(trade_data.get("shares", 0), 2),
                    round(stake, 2),
                    round(pnl, 2),
                    f"{pnl_pct:.1f}%",
                    trade_data.get("close_reason", ""),
                    round(trade_data.get("duration", 0), 1),
                    round(self._balance, 2),
                ]

                ws.append(row_data)
                row_num = ws.max_row

                fill = GREEN_FILL if pnl >= 0 else RED_FILL
                for col in range(1, len(TRADE_HEADERS) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="center")

                # Track daily data
                if date_str not in self._daily_data:
                    self._daily_data[date_str] = {
                        "trades": 0, "wins": 0, "losses": 0,
                        "pnl": 0, "volume": 0
                    }
                day = self._daily_data[date_str]
                day["trades"] += 1
                day["pnl"] += pnl
                day["volume"] += stake
                if pnl >= 0:
                    day["wins"] += 1
                else:
                    day["losses"] += 1

                wb.save(self.file_path)
                wb.close()
            except Exception as e:
                print(f"Backtest log error: {e}")

    def _write_event(self, event_type: str, details: str):
        """Write event to Events sheet."""
        with self._lock:
            try:
                wb = load_workbook(self.file_path)
                ws = wb["Events"]
                now = datetime.now()
                ws.append([
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S"),
                    event_type,
                    details[:200],
                ])
                row_num = ws.max_row
                for col in range(1, len(EVENT_HEADERS) + 1):
                    ws.cell(row=row_num, column=col).border = BORDER
                wb.save(self.file_path)
                wb.close()
            except Exception as e:
                print(f"Backtest event error: {e}")

    def finalize(self):
        """Finalize the backtest — update Summary and Daily sheets."""
        with self._lock:
            try:
                wb = load_workbook(self.file_path)
                ws_trades = wb["Trades"]
                ws_summary = wb["Summary"]
                ws_daily = wb["Daily"]

                # Collect all P&L values and balance curve
                pnls = []
                volumes = []
                durations = []
                balance_curve = [self._starting_balance]

                for row in ws_trades.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    try:
                        pnl = float(row[10])
                        pnls.append(pnl)
                        balance_curve.append(balance_curve[-1] + pnl)
                    except (ValueError, TypeError):
                        pass
                    try:
                        volumes.append(float(row[9]))
                    except (ValueError, TypeError):
                        pass
                    try:
                        durations.append(float(row[13]))
                    except (ValueError, TypeError):
                        pass

                total = len(pnls)
                wins = sum(1 for p in pnls if p >= 0)
                losses = sum(1 for p in pnls if p < 0)
                win_rate = (wins / total * 100) if total > 0 else 0
                total_pnl = sum(pnls)
                best = max(pnls) if pnls else 0
                worst = min(pnls) if pnls else 0
                avg_pnl = total_pnl / total if total > 0 else 0
                total_vol = sum(volumes)
                avg_dur = sum(durations) / len(durations) if durations else 0

                # Calculate max drawdown
                peak = self._starting_balance
                max_dd = 0
                max_dd_pct = 0
                for bal in balance_curve:
                    if bal > peak:
                        peak = bal
                    dd = peak - bal
                    if dd > max_dd:
                        max_dd = dd
                        max_dd_pct = (dd / peak * 100) if peak > 0 else 0

                # Profit factor
                gross_profit = sum(p for p in pnls if p > 0)
                gross_loss = abs(sum(p for p in pnls if p < 0))
                profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float('inf')

                return_pct = ((self._balance - self._starting_balance) / self._starting_balance * 100) if self._starting_balance > 0 else 0

                first_date = ws_trades.cell(row=2, column=2).value if total > 0 else "N/A"
                last_date = ws_trades.cell(row=ws_trades.max_row, column=2).value if total > 0 else "N/A"

                # Update summary
                summary_values = [
                    self._starting_balance,
                    round(self._balance, 2),
                    round(total_pnl, 2),
                    f"{return_pct:.1f}%",
                    total,
                    wins,
                    losses,
                    f"{win_rate:.1f}%",
                    round(best, 2),
                    round(worst, 2),
                    round(avg_pnl, 2),
                    round(max_dd, 2),
                    f"{max_dd_pct:.1f}%",
                    f"{profit_factor:.2f}" if profit_factor != float('inf') else "∞",
                    round(total_vol, 2),
                    round(avg_dur, 1),
                    first_date,
                    last_date,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ]
                for i, val in enumerate(summary_values):
                    ws_summary.cell(row=i + 2, column=2, value=val)

                # Write daily data
                running_bal = self._starting_balance
                for date_str in sorted(self._daily_data.keys()):
                    day = self._daily_data[date_str]
                    running_bal += day["pnl"]
                    wr = (day["wins"] / day["trades"] * 100) if day["trades"] > 0 else 0
                    ws_daily.append([
                        date_str,
                        day["trades"],
                        day["wins"],
                        day["losses"],
                        f"{wr:.0f}%",
                        round(day["pnl"], 2),
                        round(day["volume"], 2),
                        round(running_bal, 2),
                    ])
                    row_num = ws_daily.max_row
                    fill = GREEN_FILL if day["pnl"] >= 0 else RED_FILL
                    for col in range(1, len(DAILY_HEADERS) + 1):
                        cell = ws_daily.cell(row=row_num, column=col)
                        cell.fill = fill
                        cell.border = BORDER
                        cell.alignment = Alignment(horizontal="center")

                # Write BACKTEST_COMPLETE event directly (avoid lock deadlock)
                ws_events = wb["Events"]
                now_fin = datetime.now()
                ws_events.append([
                    now_fin.strftime("%Y-%m-%d"),
                    now_fin.strftime("%H:%M:%S"),
                    "BACKTEST_COMPLETE",
                    f"Trades: {total} | Win Rate: {win_rate:.1f}% | "
                    f"P&L: ${total_pnl:+.2f} | Return: {return_pct:+.1f}%",
                ])

                wb.save(self.file_path)
                wb.close()

                return {
                    "total": total, "wins": wins, "losses": losses,
                    "win_rate": win_rate, "total_pnl": total_pnl,
                    "return_pct": return_pct, "best": best, "worst": worst,
                    "max_drawdown": max_dd, "max_drawdown_pct": max_dd_pct,
                    "profit_factor": profit_factor, "final_balance": self._balance,
                }
            except Exception as e:
                print(f"Backtest finalize error: {e}")
                return {}

    def get_summary_text(self) -> str:
        """Get a formatted summary for Telegram."""
        result = self.finalize()
        if not result:
            return "No backtest data available."

        s = result
        pnl_sign = "+" if s["total_pnl"] >= 0 else ""
        pnl_emoji = "🟢" if s["total_pnl"] >= 0 else "🔴"

        return (
            f"📊 <b>BACKTEST RESULTS</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"💰 <b>Balance:</b>\n"
            f"Start: ${self._starting_balance:.2f}\n"
            f"Final: <b>${s['final_balance']:.2f}</b>\n"
            f"Return: <b>{pnl_sign}{s['return_pct']:.1f}%</b>\n\n"
            f"📈 <b>Performance:</b>\n"
            f"Total Trades: {s['total']}\n"
            f"Wins: {s['wins']} | Losses: {s['losses']}\n"
            f"Win Rate: <b>{s['win_rate']:.1f}%</b>\n\n"
            f"{pnl_emoji} <b>P&L:</b>\n"
            f"Total: <b>{pnl_sign}${s['total_pnl']:.2f}</b>\n"
            f"Best: +${s['best']:.2f}\n"
            f"Worst: ${s['worst']:.2f}\n\n"
            f"📉 <b>Risk:</b>\n"
            f"Max Drawdown: ${s['max_drawdown']:.2f} ({s['max_drawdown_pct']:.1f}%)\n"
            f"Profit Factor: {s['profit_factor']:.2f}\n\n"
            f"📁 Full data: backtest_log.xlsx"
        )
