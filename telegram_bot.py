"""
╔══════════════════════════════════════════════════════════════╗
║   POLYMARKET TELEGRAM BOT — TELEGRAM INTEGRATION v3.0       ║
╚══════════════════════════════════════════════════════════════╝
Manual trading + full parameter control via Telegram.

Trading Commands:
  /buy up [timeframe]      - Buy UP tokens
  /buy down [timeframe]    - Buy DOWN tokens
  /sell                    - Sell current position
  /auto up                 - Auto-repeat buying UP
  /auto down               - Auto-repeat buying DOWN
  /auto off                - Stop auto-repeat

Status Commands:
  /status   - Bot status, open positions, daily P&L
  /config   - Show ALL current parameters
  /pnl      - P&L summary
  /trades   - Recent trades
  /markets  - Available markets
  /balance  - Wallet balance

Parameter Commands:
  /set tp <value>          - Take-profit %
  /set sl <value>          - Stop-loss %
  /set amount <value>      - Trade amount ($)
  /set percent <value>     - Portfolio %
  /set size fixed|percent  - Sizing mode
  /set slippage <value>    - Max slippage ($)
  /set shareprice <value>  - Target share price
  /set market <5m,15m,...> - Market timeframes
  /set maxtrades <value>   - Max trades per day
  /set cooldown <value>    - Cooldown minutes
  /set tick <value>        - Tick interval (seconds)

Control:
  /start  - Resume bot
  /stop   - Pause bot
  /help   - All commands
"""

import threading
import time
from datetime import datetime, timezone
from typing import Optional

import requests

from config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, trading_config


class TelegramNotifier:
    """Sends notifications to Telegram chat."""

    def __init__(self):
        self.token = TELEGRAM_BOT_TOKEN
        self.chat_id = TELEGRAM_CHAT_ID
        self.base_url = f"https://api.telegram.org/bot{self.token}"
        self._enabled = bool(
            self.token and self.chat_id
            and self.token != "your_telegram_bot_token_here"
            and self.chat_id != "your_chat_id_here"
        )

    @property
    def is_enabled(self) -> bool:
        return self._enabled

    def send(self, message: str, parse_mode: str = "HTML") -> bool:
        """Send a message to Telegram."""
        if not self._enabled:
            return False
        try:
            resp = requests.post(
                f"{self.base_url}/sendMessage",
                json={
                    "chat_id": self.chat_id,
                    "text": message,
                    "parse_mode": parse_mode,
                    "disable_web_page_preview": True,
                },
                timeout=10,
            )
            return resp.status_code == 200
        except Exception as e:
            print(f"Telegram send error: {e}")
            return False

    def send_document(self, file_path: str, caption: str = "") -> bool:
        """Send a document/file to Telegram."""
        import os
        if not self._enabled or not os.path.exists(file_path):
            return False
        try:
            with open(file_path, "rb") as f:
                resp = requests.post(
                    f"{self.base_url}/sendDocument",
                    data={"chat_id": self.chat_id, "caption": caption},
                    files={"document": f},
                    timeout=30,
                )
            return resp.status_code == 200
        except Exception as e:
            print(f"Telegram send document error: {e}")
            return False

    def send_trade_opened(self, trade) -> bool:
        direction = "🟢 LONG (UP)" if trade.direction.value == "UP" else "🔴 SHORT (DOWN)"
        msg = (
            f"📈 <b>TRADE OPENED</b>\n\n"
            f"Direction: {direction}\n"
            f"Market: {trade.timeframe} timeframe\n"
            f"Stake: <b>${trade.amount:.2f}</b>\n"
            f"Entry Price: ${trade.share_price:.4f}/share\n"
            f"Shares: {trade.shares:.1f}\n"
            f"Order ID: <code>{trade.order_id[:20]}...</code>\n"
            f"Time: {trade.entry_time} UTC"
        )
        return self.send(msg)

    def send_trade_closed(self, trade) -> bool:
        pnl_emoji = "💰" if trade.pnl >= 0 else "💸"
        pnl_sign = "+" if trade.pnl >= 0 else ""
        direction = "🟢 UP" if trade.direction.value == "UP" else "🔴 DOWN"

        reason_map = {
            "TAKE_PROFIT|SOLD": "🎯 Take-Profit (Sold ✅)",
            "TAKE_PROFIT|SELL_FAILED": "🎯 Take-Profit (Sell ⚠️)",
            "STOP_LOSS|SOLD": "🛑 Stop-Loss (Sold ✅)",
            "STOP_LOSS|SELL_FAILED": "🛑 Stop-Loss (Sell ⚠️)",
            "MANUAL_SELL|SOLD": "📤 Manual Sell (Sold ✅)",
            "MANUAL_SELL|SELL_FAILED": "📤 Manual Sell (Sell ⚠️)",
            "MARKET_WIN": "✅ Market Resolved (WIN)",
            "MARKET_LOSS": "❌ Market Resolved (LOSS)",
            "WIN": "✅ WIN",
            "LOSS": "❌ LOSS",
            "CANCELLED": "🚫 Cancelled",
        }
        reason_text = reason_map.get(trade.close_reason, trade.close_reason or "Closed")

        msg = (
            f"{pnl_emoji} <b>TRADE CLOSED</b>\n\n"
            f"Direction: {direction}\n"
            f"Reason: {reason_text}\n"
            f"Entry Price: ${trade.share_price:.4f}\n"
            f"Exit Price: ${trade.result_price:.4f}\n"
            f"P&L: <b>{pnl_sign}${trade.pnl:.2f}</b>\n"
            f"Market: {trade.timeframe} timeframe"
        )
        return self.send(msg)

    def send_status(self, trader, engine_state=None) -> bool:
        from trade_manager import TradeStatus

        status = "🟢 RUNNING" if trading_config.bot_running else "🔴 PAUSED"
        state_text = ""
        if engine_state:
            state_text = f"\nState: {engine_state.bot_state.value}"
            if engine_state.auto_repeat_active:
                dir_text = engine_state.auto_repeat_direction.value if engine_state.auto_repeat_direction else "?"
                state_text += f"\n🔄 Auto-Repeat: {dir_text}"

        open_trade_text = "None"
        if trader.current_trade and trader.current_trade.status == TradeStatus.OPEN:
            t = trader.current_trade
            dir_text = "UP 🟢" if t.direction.value == "UP" else "DOWN 🔴"
            upnl = t.unrealized_pnl
            upnl_sign = "+" if upnl >= 0 else ""
            open_trade_text = (
                f"\n  Direction: {dir_text}\n"
                f"  Stake: ${t.amount:.2f}\n"
                f"  Entry: ${t.share_price:.4f}\n"
                f"  Current: ${t.current_price:.4f}\n"
                f"  Unrealized P&L: {upnl_sign}${upnl:.2f} ({t.unrealized_pnl_pct:+.1f}%)"
            )

        msg = (
            f"📊 <b>BOT STATUS</b>\n\n"
            f"Status: {status}{state_text}\n\n"
            f"── Martingale Recovery ──\n"
            f"Loss Streak: {trader.loss_streak}\n"
            f"Next Stake: <b>${trader.get_trade_amount():.2f}</b>\n\n"
            f"── Open Position ──\n{open_trade_text}\n\n"
            f"── Today ──\n"
            f"Daily P&L: <b>${trader.daily_pnl:+.2f}</b>\n"
            f"Trades Today: {engine_state.trades_today if engine_state else '?'}\n\n"
            f"── All Time ──\n"
            f"Total Trades: {trader.total_trades}\n"
            f"Win Rate: {trader.win_rate:.1f}%\n"
            f"Wins: {trader.wins} | Losses: {trader.losses}\n"
            f"Total P&L: <b>${trader.total_pnl:+.2f}</b>\n"
            f"Volume: ${trader.total_volume:.2f}"
        )
        return self.send(msg)

    def send_config(self) -> bool:
        cfg = trading_config
        size_text = (
            f"${cfg.trade_amount:.2f} (fixed)"
            if cfg.trade_size_mode == "fixed"
            else f"{cfg.trade_percent:.1f}% of portfolio"
        )
        auto_rpt = "ON" if cfg.auto_repeat else "OFF"

        msg = (
            f"⚙️ <b>ALL PARAMETERS</b>\n\n"
            f"── Trading ──\n"
            f"Trade Size: {size_text}\n"
            f"Take-Profit: {cfg.take_profit_pct:.1f}%\n"
            f"Stop-Loss: {cfg.stop_loss_pct:.1f}%\n"
            f"Share Price: ${cfg.share_price:.2f}\n"
            f"Max Slippage: ${cfg.max_slippage:.3f}\n"
            f"Loss Multiplier: {cfg.loss_multiplier:.1f}% (Martingale)\n\n"
            f"── Markets ──\n"
            f"Timeframes: {', '.join(cfg.market_timeframes)}\n\n"
            f"── Auto-Repeat ──\n"
            f"Auto-Repeat: {auto_rpt}\n"
            f"Max Trades/Day: {cfg.max_trades_per_day}\n\n"
            f"── Timing ──\n"
            f"Cooldown: {cfg.cooldown_minutes} min\n"
            f"Tick Interval: {cfg.tick_interval}s\n"
            f"Max Entry Wait: {cfg.max_entry_wait_minutes} min"
        )
        return self.send(msg)

    def send_pnl_summary(self, trader) -> bool:
        pnl_emoji = "📈" if trader.total_pnl >= 0 else "📉"
        daily_emoji = "📈" if trader.daily_pnl >= 0 else "📉"
        msg = (
            f"{pnl_emoji} <b>P&L SUMMARY</b>\n\n"
            f"── Today ──\n"
            f"{daily_emoji} Daily P&L: <b>${trader.daily_pnl:+.2f}</b>\n\n"
            f"── All Time ──\n"
            f"Total P&L: <b>${trader.total_pnl:+.2f}</b>\n"
            f"Total Volume: ${trader.total_volume:.2f}\n"
            f"Total Trades: {trader.total_trades}\n"
            f"Win Rate: {trader.win_rate:.1f}%\n"
            f"Wins: {trader.wins} | Losses: {trader.losses}"
        )
        return self.send(msg)

    def send_recent_trades(self, trader) -> bool:
        trades = trader.recent_trades
        if not trades:
            return self.send("📜 <b>RECENT TRADES</b>\n\nNo trades yet.")
        lines = ["📜 <b>RECENT TRADES (Last 10)</b>\n"]
        for i, t in enumerate(trades, 1):
            dir_icon = "🟢" if t.direction.value == "UP" else "🔴"
            pnl_sign = "+" if t.pnl >= 0 else ""
            lines.append(
                f"{i}. {dir_icon} {t.entry_time} | "
                f"${t.amount:.2f} | "
                f"{pnl_sign}${t.pnl:.2f} | "
                f"{t.close_reason[:15]}"
            )
        return self.send("\n".join(lines))

    def send_error(self, error_msg: str) -> bool:
        msg = f"⚠️ <b>BOT ERROR</b>\n\n<code>{error_msg[:500]}</code>"
        return self.send(msg)

    def send_bot_started(self) -> bool:
        mode = "🔴 PAPER MODE" if trading_config.paper_mode else "🟢 LIVE TRADING"
        cfg = trading_config
        msg = (
            f"🚀 <b>BOT STARTED</b>\n\n"
            f"Mode: {mode}\n"
            f"Markets: {', '.join(cfg.market_timeframes)}\n"
            f"TP: {cfg.take_profit_pct}% | SL: {cfg.stop_loss_pct}%\n"
            f"Trade Size: ${cfg.trade_amount} | Slippage: ${cfg.max_slippage}\n\n"
            f"<b>Commands:</b>\n"
            f"/buy up — Buy UP tokens\n"
            f"/buy down — Buy DOWN tokens\n"
            f"/sell — Sell current position\n"
            f"/auto up — Auto-repeat UP\n"
            f"/help — All commands\n\n"
            f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC"
        )
        return self.send(msg)

    def send_bot_stopped(self) -> bool:
        msg = f"🛑 <b>BOT STOPPED</b>\n\nTime: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC"
        return self.send(msg)


class TelegramCommandHandler:
    """Polls Telegram for commands and executes them."""

    def __init__(self, notifier: TelegramNotifier, trader=None, engine=None, market_finder=None):
        self.notifier = notifier
        self.trader = trader
        self.engine = engine
        self.market_finder = market_finder
        self._last_update_id = 0
        self._running = False
        self._thread: Optional[threading.Thread] = None
        self._search_results: list = []  # Cache for /trade command

    def start(self):
        if not self.notifier.is_enabled:
            return
        # Flush any old pending updates before starting
        self._flush_old_updates()
        # Register commands in Telegram menu
        self._register_commands_menu()
        self._running = True
        self._thread = threading.Thread(target=self._poll_loop, daemon=True)
        self._thread.start()

    def stop(self):
        self._running = False
        if self._thread:
            self._thread.join(timeout=5)

    def _flush_old_updates(self):
        """Flush all pending Telegram updates so we don't re-process old commands."""
        try:
            resp = requests.get(
                f"{self.notifier.base_url}/getUpdates",
                params={"offset": -1, "timeout": 0},
                timeout=10,
            )
            if resp.status_code == 200:
                data = resp.json()
                results = data.get("result", [])
                if results:
                    self._last_update_id = results[-1]["update_id"]
        except Exception:
            pass

    def _poll_loop(self):
        while self._running:
            try:
                self._poll_updates()
            except Exception as e:
                print(f"Telegram poll error: {e}")
            time.sleep(1)

    def _poll_updates(self):
        try:
            resp = requests.get(
                f"{self.notifier.base_url}/getUpdates",
                params={
                    "offset": self._last_update_id + 1,
                    "timeout": 5,
                    "allowed_updates": '["message"]',
                },
                timeout=15,
            )
            if resp.status_code != 200:
                return
            data = resp.json()
            if not data.get("ok"):
                return
            for update in data.get("result", []):
                self._last_update_id = update["update_id"]
                message = update.get("message", {})
                chat_id = str(message.get("chat", {}).get("id", ""))
                if chat_id != self.notifier.chat_id:
                    continue

                # Ignore messages older than 30 seconds
                msg_date = message.get("date", 0)
                if msg_date > 0:
                    import time as _time
                    age = _time.time() - msg_date
                    if age > 30:
                        continue

                text = message.get("text", "").strip()
                if text.startswith("/"):
                    self._handle_command(text)
        except Exception:
            pass

    def _handle_command(self, text: str):
        parts = text.split()
        cmd = parts[0].lower().split("@")[0]
        args = parts[1:]

        handlers = {
            "/buy": self._cmd_buy,
            "/sell": self._cmd_sell,
            "/auto": self._cmd_auto,
            "/schedule": self._cmd_schedule,
            "/search": self._cmd_search,
            "/trending": self._cmd_trending,
            "/trade": self._cmd_trade,
            "/status": self._cmd_status,
            "/config": self._cmd_config,
            "/set": self._cmd_set,
            "/start": self._cmd_start,
            "/stop": self._cmd_stop,
            "/pnl": self._cmd_pnl,
            "/trades": self._cmd_trades,
            "/history": self._cmd_history,
            "/backtest": self._cmd_backtest,
            "/export": self._cmd_export,
            "/markets": self._cmd_markets,
            "/balance": self._cmd_balance,
            "/help": self._cmd_help,
        }

        handler = handlers.get(cmd)
        if handler:
            try:
                handler(args)
            except Exception as e:
                self.notifier.send(f"⚠️ Command error: {str(e)[:200]}")
        else:
            self.notifier.send(f"❓ Unknown: {cmd}\nType /help for commands.")

    # ── Trading Commands ─────────────────────────────

    def _cmd_buy(self, args):
        """Handle /buy up [timeframe] or /buy down [timeframe]"""
        if not args:
            self.notifier.send(
                "🛒 <b>BUY — BTC Markets</b>\n\n"
                "<b>BTC UP/DOWN:</b>\n"
                "/buy up — BTC will go UP\n"
                "/buy down — BTC will go DOWN\n"
                "/buy up 5m — 5-minute market\n"
                "/buy down 1h — 1-hour market\n\n"
                "<b>Custom Markets:</b>\n"
                "/search trump — Find any market\n"
                "/trending — See popular markets\n"
                "/trade 1 yes — Buy from search results"
            )
            return

        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        direction_str = args[0].lower()
        if direction_str not in ("up", "down", "yes", "no"):
            self.notifier.send("❌ Use: /buy up or /buy down\nFor custom markets use /search + /trade")
            return

        from trade_manager import TradeDirection
        # Map yes/no to up/down
        if direction_str in ("up", "yes"):
            direction = TradeDirection.UP
        else:
            direction = TradeDirection.DOWN

        # Optional timeframe
        timeframe = args[1] if len(args) > 1 else None
        if timeframe and timeframe not in ("5m", "15m", "1h", "1d"):
            self.notifier.send(f"❌ Invalid timeframe: {timeframe}. Use: 5m, 15m, 1h, 1d")
            return

        success, msg = self.engine.manual_buy(direction, timeframe)
        self.notifier.send(msg)

    def _cmd_sell(self, args):
        """Handle /sell"""
        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        success, msg = self.engine.manual_sell()
        self.notifier.send(msg)

    def _cmd_auto(self, args):
        """Handle /auto up, /auto down, /auto off"""
        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        if not args:
            self.notifier.send(
                "🔄 <b>AUTO-REPEAT</b>\n\n"
                "Usage:\n"
                "/auto up — Automatically re-buy UP after each market\n"
                "/auto down — Automatically re-buy DOWN after each market\n"
                "/auto off — Turn off auto-repeat\n\n"
                "When enabled, bot will automatically place the same\n"
                "direction trade on each new market after the previous\n"
                "one resolves."
            )
            return

        action = args[0].lower()
        if action == "up":
            from trade_manager import TradeDirection
            msg = self.engine.enable_auto_repeat(TradeDirection.UP)
            # Also auto-buy if currently idle
            if self.engine.state.bot_state.value == "IDLE" and not self.trader.has_open_trade():
                success, buy_msg = self.engine.manual_buy(TradeDirection.UP)
                if success:
                    msg += f"\n\n{buy_msg}"
            self.notifier.send(msg)
        elif action == "down":
            from trade_manager import TradeDirection
            msg = self.engine.enable_auto_repeat(TradeDirection.DOWN)
            if self.engine.state.bot_state.value == "IDLE" and not self.trader.has_open_trade():
                success, buy_msg = self.engine.manual_buy(TradeDirection.DOWN)
                if success:
                    msg += f"\n\n{buy_msg}"
            self.notifier.send(msg)
        elif action == "off":
            msg = self.engine.disable_auto_repeat()
            self.notifier.send(msg)
        else:
            self.notifier.send("❌ Use: /auto up, /auto down, or /auto off")

    def _cmd_schedule(self, args):
        """Handle /schedule HH:MM up/down or /schedule off"""
        if not args:
            # Show current schedule if any
            if self.engine and self.engine.state.scheduled_time:
                st = self.engine.state
                dir_text = st.scheduled_direction.value if st.scheduled_direction else "?"
                self.notifier.send(
                    f"⏰ <b>SCHEDULED TRADE</b>\n\n"
                    f"Time: <b>{st.scheduled_time}</b>\n"
                    f"Direction: {dir_text}\n"
                    f"Market: {st.scheduled_timeframe or '?'}\n\n"
                    f"Cancel with /schedule off"
                )
            else:
                self.notifier.send(
                    "⏰ <b>SCHEDULE A TRADE</b>\n\n"
                    "Usage:\n"
                    "/schedule 14:30 up — Buy UP at 14:30\n"
                    "/schedule 2:24 down — Buy DOWN at 2:24\n"
                    "/schedule 9:00 up 5m — Buy UP at 9:00 on 5m market\n"
                    "/schedule off — Cancel scheduled trade\n\n"
                    "Time is in your local timezone (24h format)."
                )
            return

        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        # Cancel
        if args[0].lower() == "off":
            msg = self.engine.cancel_schedule()
            self.notifier.send(msg)
            return

        # Need at least time + direction
        if len(args) < 2:
            self.notifier.send("❌ Usage: /schedule HH:MM up/down\nExample: /schedule 14:30 up")
            return

        time_str = args[0]
        direction_str = args[1].lower()

        if direction_str not in ("up", "down"):
            self.notifier.send("❌ Direction must be 'up' or 'down'")
            return

        from trade_manager import TradeDirection
        direction = TradeDirection.UP if direction_str == "up" else TradeDirection.DOWN

        # Optional timeframe
        timeframe = args[2] if len(args) > 2 else None
        if timeframe and timeframe not in ("5m", "15m", "1h", "1d"):
            self.notifier.send(f"❌ Invalid timeframe: {timeframe}. Use: 5m, 15m, 1h, 1d")
            return

        success, msg = self.engine.schedule_trade(time_str, direction, timeframe)
        self.notifier.send(msg)

    # ── Custom Market Commands ────────────────────────

    def _cmd_search(self, args):
        """Handle /search <query>"""
        if not args:
            self.notifier.send(
                "🔍 <b>SEARCH MARKETS</b>\n\n"
                "Usage:\n"
                "/search bitcoin — BTC markets\n"
                "/search trump — Trump markets\n"
                "/search election — Election markets\n"
                "/search ethereum — ETH markets\n\n"
                "Then use /trade <#> <outcome> to buy!"
            )
            return

        if not self.market_finder:
            self.notifier.send("⚠️ Market finder not initialized.")
            return

        query = " ".join(args)
        self.notifier.send(f"🔍 Searching for '{query}'...")

        results = self.market_finder.search_markets(query, limit=8)
        if not results:
            self.notifier.send(f"❌ No active markets found for '{query}'")
            return

        self._search_results = results
        lines = [f"🔍 <b>SEARCH: '{query}'</b>\n"]
        for i, event in enumerate(results, 1):
            outcomes_text = event.outcome_summary
            liq_text = f"${event.liquidity:,.0f}" if event.liquidity > 0 else "N/A"
            lines.append(
                f"<b>{i}.</b> {event.question[:55]}\n"
                f"   {outcomes_text}\n"
                f"   💧 Liquidity: {liq_text}\n"
            )
        lines.append("\n<b>To trade:</b> /trade <i>#</i> <i>outcome</i>")
        lines.append("Example: /trade 1 yes")
        self.notifier.send("\n".join(lines))

    def _cmd_trending(self, args):
        """Handle /trending"""
        if not self.market_finder:
            self.notifier.send("⚠️ Market finder not initialized.")
            return

        self.notifier.send("📈 Fetching trending markets...")

        results = self.market_finder.get_trending_markets(limit=8)
        if not results:
            self.notifier.send("❌ Could not fetch trending markets.")
            return

        self._search_results = results
        lines = ["📈 <b>TRENDING MARKETS</b>\n"]
        for i, event in enumerate(results, 1):
            outcomes_text = event.outcome_summary
            liq_text = f"${event.liquidity:,.0f}" if event.liquidity > 0 else "N/A"
            lines.append(
                f"<b>{i}.</b> {event.question[:55]}\n"
                f"   {outcomes_text}\n"
                f"   💧 Liquidity: {liq_text}\n"
            )
        lines.append("\n<b>To trade:</b> /trade <i>#</i> <i>outcome</i>")
        lines.append("Example: /trade 1 yes")
        self.notifier.send("\n".join(lines))

    def _cmd_trade(self, args):
        """Handle /trade <number> <outcome>"""
        if not args or len(args) < 2:
            self.notifier.send(
                "📋 <b>TRADE CUSTOM MARKET</b>\n\n"
                "Usage:\n"
                "/trade 1 yes — Buy 'Yes' on result #1\n"
                "/trade 1 no — Buy 'No' on result #1\n"
                "/trade 3 1 — Buy outcome #1 of result #3\n\n"
                "First use /search or /trending to see markets."
            )
            return

        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        if not self._search_results:
            self.notifier.send("⚠️ No search results. Use /search or /trending first.")
            return

        try:
            index = int(args[0]) - 1
        except ValueError:
            self.notifier.send("❌ First argument must be a number (e.g., /trade 1 yes)")
            return

        if index < 0 or index >= len(self._search_results):
            self.notifier.send(f"❌ Invalid #. Choose 1-{len(self._search_results)}")
            return

        event = self._search_results[index]

        # Refresh prices
        event = self.market_finder.refresh_event_prices(event)

        # Determine outcome
        outcome_str = args[1].lower()
        outcome_index = None

        # Try matching by name
        for i, outcome in enumerate(event.outcomes):
            if outcome.lower() == outcome_str:
                outcome_index = i
                break

        # Try matching yes/no
        if outcome_index is None:
            if outcome_str in ("yes", "up", "1", "first"):
                outcome_index = 0
            elif outcome_str in ("no", "down", "2", "second"):
                outcome_index = 1
            else:
                try:
                    outcome_index = int(outcome_str) - 1
                except ValueError:
                    pass

        if outcome_index is None or outcome_index < 0 or outcome_index >= len(event.outcomes):
            outcomes_list = ", ".join(f"'{o}'" for o in event.outcomes)
            self.notifier.send(f"❌ Invalid outcome: {outcome_str}\nAvailable: {outcomes_list}")
            return

        # Confirm and trade
        outcome_name = event.outcomes[outcome_index]
        price = event.get_price_for_outcome(outcome_index)
        self.notifier.send(
            f"📋 Trading: <b>{outcome_name}</b> @ ${price:.3f}\n"
            f"Market: {event.question[:60]}\n\n"
            f"Placing order..."
        )

        success, msg = self.engine.manual_buy_custom(event, outcome_index)
        self.notifier.send(msg)

    # ── Status Commands ──────────────────────────────

    def _cmd_status(self, args):
        if self.trader:
            engine_state = self.engine.state if self.engine else None
            self.notifier.send_status(self.trader, engine_state)
        else:
            self.notifier.send("⚠️ Trader not initialized yet.")

    def _cmd_config(self, args):
        self.notifier.send_config()

    def _cmd_pnl(self, args):
        if self.trader:
            self.notifier.send_pnl_summary(self.trader)
        else:
            self.notifier.send("⚠️ Trader not initialized yet.")

    def _cmd_trades(self, args):
        if self.trader:
            self.notifier.send_recent_trades(self.trader)
        else:
            self.notifier.send("⚠️ Trader not initialized yet.")

    def _cmd_history(self, args):
        """Handle /history [period] — show historical trade data."""
        if not self.engine:
            self.notifier.send("⚠️ Engine not initialized yet.")
            return

        period = args[0] if args else None

        # Show usage if they just type /history
        if not period:
            self.notifier.send(
                "📊 <b>TRADE HISTORY</b>\n\n"
                "Usage:\n"
                "/history all \u2014 All time\n"
                "/history today \u2014 Today only\n"
                "/history 7d \u2014 Last 7 days\n"
                "/history 1m \u2014 Last 1 month\n"
                "/history 3m \u2014 Last 3 months\n"
                "/history 1y \u2014 Last 1 year\n"
                "/history 2y \u2014 Last 2 years\n"
                "/history 2024 \u2014 Year 2024\n"
                "/history 2024-03 \u2014 March 2024"
            )
            return

        data = self.engine.logger.get_history(period)
        trades = data.get("trades", [])
        stats = data.get("stats", {})
        label = data.get("period_label", "Unknown")

        if not stats or stats.get("total", 0) == 0:
            self.notifier.send(f"📊 <b>{label}</b>\n\nNo trades found for this period.")
            return

        # Build summary
        s = stats
        pnl_sign = '+' if s['total_pnl'] >= 0 else ''
        pnl_emoji = '🟢' if s['total_pnl'] >= 0 else '🔴'

        msg = (
            f"📊 <b>HISTORY: {label}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📈 <b>Summary:</b>\n"
            f"Total Trades: <b>{s['total']}</b>\n"
            f"Wins: {s['wins']} | Losses: {s['losses']}\n"
            f"Win Rate: <b>{s['win_rate']:.1f}%</b>\n\n"
            f"{pnl_emoji} <b>P&L:</b>\n"
            f"Total: <b>{pnl_sign}${s['total_pnl']:.2f}</b>\n"
            f"Avg per Trade: ${s['avg_pnl']:.2f}\n"
            f"Best: +${s['best']:.2f}\n"
            f"Worst: ${s['worst']:.2f}\n\n"
            f"💰 Volume: ${s['total_volume']:.2f}\n"
            f"📅 {s['first_date']} \u2192 {s['last_date']}"
        )

        # Show last 10 trades of the period
        if trades:
            recent = trades[-10:]
            msg += "\n\n<b>Recent Trades:</b>\n"
            for t in recent:
                pnl_val = float(t['pnl']) if t['pnl'] else 0
                icon = '✅' if pnl_val >= 0 else '❌'
                sign = '+' if pnl_val >= 0 else ''
                msg += (
                    f"{icon} {t['date']} {t['time'][:5]} | "
                    f"{t['direction']} | {sign}${pnl_val:.2f} | "
                    f"{t['close_reason']}\n"
                )

        self.notifier.send(msg)

    def _cmd_backtest(self, args):
        """Handle /backtest — run a quick simulation and save to backtest_log.xlsx."""
        if not args:
            self.notifier.send(
                "🧪 <b>BACKTEST</b>\n\n"
                "Usage:\n"
                "/backtest run — Run backtest with live trade history\n"
                "/backtest results — Show last backtest results\n\n"
                "Results saved to <b>backtest_log.xlsx</b>\n"
                "Sheets: Trades, Events, Summary, Daily"
            )
            return

        action = args[0].lower()

        if action == "results":
            # Show last backtest results
            import os
            bt_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backtest_log.xlsx")
            if not os.path.exists(bt_file):
                self.notifier.send("❌ No backtest data found. Run /backtest run first.")
                return
            try:
                from backtest_logger import BacktestLogger
                bt = BacktestLogger.__new__(BacktestLogger)
                bt.file_path = bt_file
                bt._lock = __import__('threading').Lock()
                bt._starting_balance = 100
                bt._balance = 100
                bt._daily_data = {}
                bt._trade_count = 0

                from openpyxl import load_workbook
                wb = load_workbook(bt_file, read_only=True)
                ws = wb["Summary"]
                lines = ["📊 <b>LAST BACKTEST RESULTS</b>\n━━━━━━━━━━━━━━━━━━━━━━━\n"]
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    if row[0]:
                        lines.append(f"{row[0]}: <b>{row[1]}</b>")
                wb.close()
                self.notifier.send("\n".join(lines))
            except Exception as e:
                self.notifier.send(f"⚠️ Error reading backtest: {str(e)[:200]}")
            return

        if action == "run":
            if not self.engine:
                self.notifier.send("⚠️ Engine not initialized yet.")
                return

            self.notifier.send("🧪 Running backtest with your trade history...\nThis may take a moment...")

            import threading
            threading.Thread(target=self._run_backtest, daemon=True).start()
            return

        self.notifier.send("❌ Unknown. Use: /backtest run or /backtest results")

    def _run_backtest(self):
        """Run backtest in background thread."""
        try:
            from backtest_logger import BacktestLogger
            from openpyxl import load_workbook
            import os

            # Read live trade log
            live_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_log.xlsx")
            if not os.path.exists(live_file):
                self.notifier.send("❌ No trade history found. Trade first, then backtest.")
                return

            wb = load_workbook(live_file, read_only=True)
            ws = wb["Trades"]

            trades = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                
                # Safely get values with default fallbacks to prevent IndexError on old logs
                def get_val(idx, default):
                    return row[idx] if len(row) > idx and row[idx] is not None else default

                trades.append({
                    "direction": get_val(3, "UP"),
                    "market": get_val(4, ""),
                    "timeframe": get_val(5, "15m"),
                    "entry_price": float(get_val(6, 0.50)),
                    "exit_price": float(get_val(7, 0.50)),
                    "shares": float(get_val(8, 10)),
                    "stake": float(get_val(9, 5)),
                    "pnl": float(get_val(10, 0)),
                    "close_reason": get_val(12, "Unknown"),
                    "duration": float(get_val(14, 0)),
                    "date": str(get_val(1, "")),
                    "time": str(get_val(2, "")),
                })
            wb.close()

            if not trades:
                self.notifier.send("❌ No trades in history. Make some trades first!")
                return

            # Calculate starting balance from config
            from config import trading_config
            starting_bal = trading_config.trade_amount * 20  # Assume 20x trade size as starting

            # Create backtest log
            bt = BacktestLogger(starting_balance=starting_bal)

            for t in trades:
                bt.log_trade(t)

            result = bt.finalize()
            msg = bt.get_summary_text()
            self.notifier.send(msg)

        except Exception as e:
            self.notifier.send(f"⚠️ Backtest error: {str(e)[:300]}")

    def _cmd_export(self, args):
        """Handle /export [live/backtest]"""
        if not args:
            self.notifier.send(
                "📁 <b>EXPORT LOGS</b>\n\n"
                "Usage:\n"
                "/export live — Send trade_log.xlsx\n"
                "/export backtest — Send backtest_log.xlsx"
            )
            return

        import os
        target = args[0].lower()
        if target == "live":
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trade_log.xlsx")
            label = "Live Trade Log"
        elif target == "backtest":
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backtest_log.xlsx")
            label = "Backtest Log"
        else:
            self.notifier.send("❌ Unknown target. Use: /export live or /export backtest")
            return

        if not os.path.exists(path):
            self.notifier.send(f"❌ File not found: {label}")
            return

        self.notifier.send(f"📤 Uploading {label}...")
        success = self.notifier.send_document(path, f"📊 {label} (Generated at {datetime.now().strftime('%Y-%m-%d %H:%M')})")
        if not success:
            self.notifier.send("❌ Failed to send file.")

    def _cmd_markets(self, args):
        if not self.market_finder:
            self.notifier.send("⚠️ Market finder not initialized.")
            return
        markets = self.market_finder.find_all_markets()
        lines = ["🏪 <b>AVAILABLE MARKETS</b>\n"]
        for tf, market in markets.items():
            if market:
                lines.append(
                    f"✅ <b>{tf}</b>: {market.question[:50]}\n"
                    f"   UP: ${market.price_up:.3f} | DOWN: ${market.price_down:.3f}\n"
                    f"   Closes in: {market.minutes_until_close:.0f} min"
                )
            else:
                lines.append(f"❌ <b>{tf}</b>: No active market")
        self.notifier.send("\n".join(lines))

    def _cmd_balance(self, args):
        """Show wallet balance."""
        from config import FUNDER_ADDRESS
        if trading_config.paper_mode:
            self.notifier.send("💰 Paper mode — no real wallet connected")
            return
        try:
            from web3 import Web3
            w3 = Web3(Web3.HTTPProvider("https://polygon-bor-rpc.publicnode.com"))
            wallet = Web3.to_checksum_address(FUNDER_ADDRESS)
            erc20_abi = [{"constant": True, "inputs": [{"name": "", "type": "address"}],
                "name": "balanceOf", "outputs": [{"name": "", "type": "uint256"}],
                "stateMutability": "view", "type": "function"}]

            # Native USDC (Polymarket uses this)
            usdc_native = w3.eth.contract(
                address=Web3.to_checksum_address("0x3c499c542cEF5E3811e1192ce70d8cC03d5c3359"),
                abi=erc20_abi,
            )
            usdc_native_bal = usdc_native.functions.balanceOf(wallet).call() / 1e6

            # USDC.e (bridged, older)
            usdc_bridged = w3.eth.contract(
                address=Web3.to_checksum_address("0x2791Bca1f2de4661ED88A30C99A7a9449Aa84174"),
                abi=erc20_abi,
            )
            usdc_bridged_bal = usdc_bridged.functions.balanceOf(wallet).call() / 1e6

            matic_bal = w3.eth.get_balance(wallet) / 1e18
            short_addr = f"{FUNDER_ADDRESS[:6]}...{FUNDER_ADDRESS[-4:]}"
            total_usdc = usdc_native_bal + usdc_bridged_bal
            msg = (
                f"💰 <b>WALLET BALANCE</b>\n\n"
                f"Wallet: <code>{short_addr}</code>\n"
                f"USDC: <b>${usdc_native_bal:.2f}</b>\n"
                f"USDC.e: ${usdc_bridged_bal:.2f}\n"
                f"Total: <b>${total_usdc:.2f}</b>\n"
                f"POL: {matic_bal:.4f}"
            )
            self.notifier.send(msg)
        except Exception as e:
            self.notifier.send(f"⚠️ Balance check failed: {str(e)[:200]}")

    # ── Parameter Commands ───────────────────────────

    def _cmd_set(self, args):
        """Handle /set commands for ALL parameters."""
        if len(args) < 2:
            msg = (
                "⚙️ <b>SET PARAMETERS</b>\n\n"
                "<b>Trading:</b>\n"
                "/set tp 90 — Take-profit %\n"
                "/set sl 30 — Stop-loss %\n"
                "/set amount 10 — Trade amount ($)\n"
                "/set percent 5 — Portfolio %\n"
                "/set size fixed — Fixed amount mode\n"
                "/set size percent — Portfolio % mode\n\n"
                "<b>Pricing:</b>\n"
                "/set slippage 0.05 — Max slippage ($)\n"
                "/set shareprice 0.50 — Target share price\n\n"
                "<b>Markets:</b>\n"
                "/set market 15m — Timeframe\n"
                "/set market 5m,15m — Multiple\n\n"
                "<b>Limits:</b>\n"
                "/set maxtrades 50 — Max trades/day\n"
                "/set cooldown 30 — Cooldown minutes\n"
                "/set tick 5 — Tick interval (sec)\n"
                "/set multiplier 100 — Loss Recovery % (0=off, 100=double)"
            )
            self.notifier.send(msg)
            return

        param = args[0].lower()
        value = args[1]

        try:
            if param == "tp":
                val = float(value)
                if val <= 0:
                    self.notifier.send("❌ TP must be > 0")
                    return
                changes = trading_config.update(take_profit_pct=val)
            elif param == "sl":
                val = float(value)
                if val <= 0:
                    self.notifier.send("❌ SL must be > 0")
                    return
                changes = trading_config.update(stop_loss_pct=val)
            elif param == "amount":
                val = float(value)
                if val <= 0:
                    self.notifier.send("❌ Amount must be > 0")
                    return
                changes = trading_config.update(trade_amount=val)
            elif param == "percent":
                val = float(value)
                if val <= 0 or val > 100:
                    self.notifier.send("❌ Percent must be 0-100")
                    return
                changes = trading_config.update(trade_percent=val)
            elif param == "size":
                mode = value.lower()
                if mode not in ("fixed", "percent"):
                    self.notifier.send("❌ Must be 'fixed' or 'percent'")
                    return
                changes = trading_config.update(trade_size_mode=mode)
            elif param == "slippage":
                val = float(value)
                if val < 0:
                    self.notifier.send("❌ Slippage must be >= 0")
                    return
                changes = trading_config.update(max_slippage=val)
            elif param == "shareprice":
                val = float(value)
                if val <= 0 or val >= 1:
                    self.notifier.send("❌ Share price must be 0-1 (e.g., 0.50)")
                    return
                changes = trading_config.update(share_price=val)
            elif param == "market":
                timeframes = [t.strip() for t in value.split(",")]
                valid_tfs = {"5m", "15m", "1h", "1d"}
                invalid = [t for t in timeframes if t not in valid_tfs]
                if invalid:
                    self.notifier.send(f"❌ Invalid: {invalid}. Use: {valid_tfs}")
                    return
                changes = trading_config.update(market_timeframes=timeframes)
            elif param == "maxtrades":
                val = int(value)
                if val <= 0:
                    self.notifier.send("❌ Must be > 0")
                    return
                changes = trading_config.update(max_trades_per_day=val)
            elif param == "cooldown":
                val = int(value)
                if val < 0:
                    self.notifier.send("❌ Must be >= 0")
                    return
                changes = trading_config.update(cooldown_minutes=val)
            elif param == "tick":
                val = int(value)
                if val < 1:
                    self.notifier.send("❌ Must be >= 1")
                    return
                changes = trading_config.update(tick_interval=val)
            elif param == "multiplier":
                val = float(value)
                if val < 0:
                    self.notifier.send("❌ Multiplier must be >= 0")
                    return
                changes = trading_config.update(loss_multiplier=val)
            elif param == "mode":
                mode = value.lower()
                if mode not in ("paper", "live"):
                    self.notifier.send("❌ Mode must be 'paper' or 'live'")
                    return
                is_paper = (mode == "paper")
                changes = trading_config.update(paper_mode=is_paper)
                if is_paper:
                    self.notifier.send("🔴 Switching to <b>PAPER MODE</b> (simulated)")
                else:
                    self.notifier.send("🟢 Switching to <b>LIVE TRADING</b> (real funds)")
            else:
                self.notifier.send(f"❌ Unknown parameter: {param}\nType /set for list.")
                return

            if changes:
                self.notifier.send(f"✅ Updated:\n" + "\n".join(changes))
            else:
                self.notifier.send("⚠️ No changes made.")

        except ValueError:
            self.notifier.send(f"❌ Invalid value: {value}")

    # ── Control Commands ─────────────────────────────

    def _cmd_start(self, args):
        trading_config.update(bot_running=True)
        if self.engine:
            self.engine.logger.log_event("BOT_RESUME", "Bot resumed by user command")
        self.notifier.send("▶️ <b>Bot RESUMED</b>\nReady for /buy commands.")

    def _cmd_stop(self, args):
        trading_config.update(bot_running=False)
        if self.engine:
            self.engine.logger.log_event("BOT_PAUSE", "Bot paused by user command")
        self.notifier.send("⏸️ <b>Bot PAUSED</b>\nUse /start to resume.")

    def _cmd_help(self, args):
        msg = (
            "🤖 <b>POLYMARKET BOT — ALL COMMANDS</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "<b>🛒 BTC TRADING:</b>\n"
            "/buy up — BTC will go UP\n"
            "/buy down — BTC will go DOWN\n"
            "/buy up 5m — 5-minute market\n"
            "/buy down 1h — 1-hour market\n"
            "/sell — Close current position\n\n"
            "<b>🌍 CUSTOM MARKETS:</b>\n"
            "/search <i>query</i> — Find any market\n"
            "/trending — Popular markets\n"
            "/trade <i># outcome</i> — Buy from results\n\n"
            "<b>⏰ SCHEDULED:</b>\n"
            "/schedule <i>HH:MM up/down</i> — Auto-trade at time\n"
            "  e.g. /schedule 14:30 up\n"
            "/schedule off — Cancel\n\n"
            "<b>🔄 AUTO-REPEAT:</b>\n"
            "/auto up — Keep buying UP\n"
            "/auto down — Keep buying DOWN\n"
            "/auto off — Stop auto\n\n"
            "<b>📊 STATUS & INFO:</b>\n"
            "/status — Positions + P&L\n"
            "/config — All settings\n"
            "/pnl — Profit/Loss summary\n"
            "/trades — Recent trade history\n"
            "/history <i>period</i> — Trade history by period\n"
            "  e.g. /history 1y, /history 2024\n"
            "/backtest run — Run backtest simulation\n"
            "/backtest results — Last backtest summary\n"
            "/export live — Download Live Excel log\n"
            "/export backtest — Download Backtest Excel log\n"
            "/markets — BTC markets list\n"
            "/balance — Wallet USDC + MATIC\n\n"
            "<b>⚙️ SETTINGS (all live!):</b>\n"
            "/set tp <i>90</i> — Take-profit %\n"
            "/set sl <i>30</i> — Stop-loss %\n"
            "/set amount <i>10</i> — Stake $\n"
            "/set multiplier <i>100</i> — Recovery %\n"
            "/set shareprice <i>0.50</i> — Target price\n"
            "/set slippage <i>0.05</i> — Range ($)\n"
            "/set mode <i>paper/live</i> — Toggle simulation\n"
            "/set market <i>5m,15m,1h</i>\n\n"
            "<b>🎮 CONTROL:</b>\n"
            "/start — Resume bot\n"
            "/stop — Pause bot\n"
            "/help — This guide"
        )
        self.notifier.send(msg)

    def _register_commands_menu(self):
        """Register the bot commands to the Telegram menu button."""
        if not self.notifier.is_enabled:
            return
        
        commands = [
            {"command": "status", "description": "📊 Check status, P&L and open trades"},
            {"command": "buy", "description": "🛒 Buy UP/DOWN (e.g. /buy up 15m)"},
            {"command": "sell", "description": "📤 Sell current open position"},
            {"command": "config", "description": "⚙️ Show all bot parameters"},
            {"command": "set", "description": "🔧 Update a setting (e.g. /set amount 10)"},
            {"command": "auto", "description": "🔄 Toggle auto-repeat mode"},
            {"command": "balance", "description": "💰 Check wallet USDC balance"},
            {"command": "pnl", "description": "📈 View daily/total P&L summary"},
            {"command": "history", "description": "🕒 View trade history by period"},
            {"command": "markets", "description": "🏪 List available BTC markets"},
            {"command": "help", "description": "❓ Show all available commands"},
        ]
        
        try:
            url = f"{self.notifier.base_url}/setMyCommands"
            import requests
            requests.post(url, json={"commands": commands}, timeout=10)
        except Exception:
            pass
